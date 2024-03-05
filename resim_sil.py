import os
import openpyxl
from tkinter import filedialog
from tkinter import Tk

# Klasör seçme penceresini açın
root = Tk()
root.withdraw()
klasor_yolu = filedialog.askdirectory(title="Klasör Seç")

if not klasor_yolu:
    print("Klasör seçilmedi. İşlem iptal edildi.")
else:
    excel_dosyasi_yolu = filedialog.askopenfilename(title="Excel Dosyasını Seç", filetypes=[("Excel Dosyaları", "*.xlsx")])

    if not excel_dosyasi_yolu:
        print("Excel dosyası seçilmedi. İşlem iptal edildi.")
    else:
        excel_dosyasi = openpyxl.load_workbook(excel_dosyasi_yolu)
        sayfa = excel_dosyasi.active

        dosya_adlari = [sayfa.cell(row=i, column=2).value for i in range(2, sayfa.max_row + 1) if sayfa.cell(row=i, column=2).value is not None]

        # Geri dönüşüm klasörü yolu
        geri_donusum_klasoru = os.path.join(klasor_yolu, "Geri Dönüşüm")
        if not os.path.exists(geri_donusum_klasoru):
            os.makedirs(geri_donusum_klasoru)

        klasordeki_dosyalar = os.listdir(klasor_yolu)

        for dosya_adi in klasordeki_dosyalar:
            kaynak_dosya_yolu = os.path.join(klasor_yolu, dosya_adi)
            hedef_dosya_yolu = os.path.join(geri_donusum_klasoru, dosya_adi)
            if dosya_adi not in dosya_adlari and os.path.isfile(kaynak_dosya_yolu):  # Ek dosya kontrolü
                try:
                    os.rename(kaynak_dosya_yolu, hedef_dosya_yolu)
                    print(f"{dosya_adi} geri dönüşüm klasörüne taşındı.")
                except OSError as e:
                    print(f"{dosya_adi} taşınamadı: {e}")

        excel_dosyasi.close()
