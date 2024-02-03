from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import glob
import os
from natsort import natsorted

from tkinter import filedialog
from tkinter import Tk

root = Tk()
root.withdraw()
klasor_yolu = filedialog.askdirectory(title="Klasör Seç")

wb = Workbook()
ws = wb.active

ws.cell(row=1, column=1, value="Resim")
ws.cell(row=1, column=2, value="Dosya Adı")
ws.cell(row=1, column=3, value="Boyut")

satir = 2
for dosya_adı in natsorted(os.listdir(klasor_yolu)):
    if dosya_adı.lower().endswith((".jpg", ".jpeg", ".png")):
        dosya_yolu = os.path.join(klasor_yolu, dosya_adı)
        boyut = f"{Image(dosya_yolu).width}x{Image(dosya_yolu).height}"
        img = Image(dosya_yolu)
        img.width = img.width * 0.4
        img.height = img.height * 0.4 
        ws.add_image(img, 'A' + str(satir))
        ws.row_dimensions[satir].height = img.height
        ws.column_dimensions['A'].width = img.width / 4.0

        # Hücrelere veri ekleyin
        ws.cell(row=satir, column=2, value=dosya_adı)
        ws.cell(row=satir, column=3, value=boyut)

        satir += 1

# Excel dosyasını kaydet
wb.save("Resim_Listesi.xlsx")
